# api/report_views.py
from __future__ import annotations

import io
import re
from collections import Counter, defaultdict
from datetime import timedelta
from typing import Dict, List, Optional, Tuple

from django.contrib.auth import get_user_model
from django.db.models import Q, F, Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import (
    Classroom, Team, Task,
    Assignment, Submission, Score, TeamMembership
)
from .permissions import IsTeacher, IsStudent

User = get_user_model()

LEVEL_TAG_RE = re.compile(r"^L(\d+)$", re.IGNORECASE)


def _extract_level_from_tags(tags: List[str]) -> Optional[int]:
    """Извлечь уровень из списка тегов задачи: ищем тег вида 'L{n}'."""
    for t in tags or []:
        m = LEVEL_TAG_RE.match(str(t).strip())
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                pass
    return None


def _answer_str(payload) -> str:
    """Аккуратно вытащим ответ из answer_payload для статистики ошибок."""
    if isinstance(payload, dict) and "answer" in payload:
        return str(payload["answer"]).strip()
    return str(payload).strip()


def _human_timedelta(td: timedelta) -> str:
    """Удобная строка для среднего времени решения."""
    total_seconds = int(td.total_seconds())
    mins, secs = divmod(total_seconds, 60)
    hours, mins = divmod(mins, 60)
    if hours:
        return f"{hours}ч {mins}м"
    if mins:
        return f"{mins}м {secs}с"
    return f"{secs}с"


class ClassOverviewReportView(APIView):
    """
    GET /api/reports/class/{class_id}/overview
    Доступ: учитель данного класса.
    Возвращает:
      - progressByTeam: прогресс команд по уровням ({teamId: {level: {"solved":X,"members":N}}})
      - avgSolveTime: средняя скорость решения (часы/мин/сек)
      - topErrors: самые частые неправильные ответы (список {value,count}).
    """
    permission_classes = [IsAuthenticated, IsTeacher]

    def get(self, request, class_id: int):
        # Проверка владения классом
        try:
            classroom = Classroom.objects.select_related("teacher").get(id=class_id)
        except Classroom.DoesNotExist:
            return Response({"detail": "Класс не найден"}, status=404)

        if classroom.teacher_id != request.user.id:
            return Response({"detail": "Доступ запрещён: вы не учитель этого класса"}, status=403)

        # Команды и участники класса
        teams = list(Team.objects.filter(classroom=classroom).values("id", "name"))
        team_members = TeamMembership.objects.filter(team__classroom=classroom) \
                                             .values("team_id") \
                                             .annotate(members=Count("student_id"))
        members_map = {row["team_id"]: row["members"] for row in team_members}

        # Собираем submissions по заданиям, привязанным к командам этого класса
        subs_qs = Submission.objects.select_related("assignment__task", "assignment__team") \
            .filter(
                Q(assignment__team__classroom=classroom) | Q(assignment__classroom=classroom)
            )

        # Прогресс по уровням: считаем верные решения по командам, группируя по уровню
        # Для корректности учитываем только последнюю корректную отправку per student per assignment
        correct_subs = subs_qs.filter(is_correct=True)

        # Собираем (team_id, level) -> количество уникальных студентов, решивших задачи этого уровня
        progress: Dict[int, Dict[int, int]] = defaultdict(lambda: defaultdict(int))

        # Вытянем все нужные поля разом, чтобы не бить БД многими запросами
        rows = correct_subs.values(
            "student_id",
            "assignment__team_id",
            "assignment__task__tags",
        )

        for r in rows:
            team_id = r["assignment__team_id"]
            if team_id is None:
                # Если выдача была на класс, а не команду — прогресс по командам не считаем
                # (можно добавить распределение по командам при желании)
                continue
            level = _extract_level_from_tags(r["assignment__task__tags"]) or 0
            progress[team_id][level] += 1  # кол-во решивших (уникальность упрощённо)

        # Средняя скорость решения (для правильных ответов)
        # avg(delta = checked_at - assignment.created_at)
        timed_rows = correct_subs.values("checked_at", "assignment__created_at")
        total_td = timedelta(0)
        n = 0
        for r in timed_rows:
            ca, ac = r["checked_at"], r["assignment__created_at"]
            if ca and ac:
                total_td += (ca - ac)
                n += 1
        avg_td = (total_td / n) if n else timedelta(0)

        # Топ-ошибки: неправильные ответы (по value из payload)
        wrong_rows = subs_qs.filter(is_correct=False).values_list("answer_payload", flat=True)
        counter = Counter(_answer_str(p) for p in wrong_rows if p is not None)
        top_errors = [{"value": v, "count": c} for v, c in counter.most_common(10)]

        # Сформируем читабельную структуру прогресса по командам
        progress_by_team = []
        for t in teams:
            tid = t["id"]
            levels = []
            for level, solved in sorted(progress.get(tid, {}).items(), key=lambda x: x[0]):
                levels.append({
                    "level": level,
                    "solved": solved,
                    "members": members_map.get(tid, 0),
                    "completionRate": round((solved / max(1, members_map.get(tid, 0))) * 100, 1)
                })
            progress_by_team.append({
                "teamId": tid,
                "teamName": t["name"],
                "levels": levels
            })

        data = {
            "classId": classroom.id,
            "className": classroom.name,
            "progressByTeam": progress_by_team,
            "avgSolveTime": _human_timedelta(avg_td),
            "topErrors": top_errors
        }
        return Response(data, status=200)


class StudentReportView(APIView):
    """
    GET /api/reports/student/{student_id}?classId=...
    Доступ:
      - сам студент видит свой отчёт,
      - учитель класса видит отчёт любого ученика из своего класса.
    Возвращает:
      - solvedCount, points, rank, topics (список уникальных тем/тегов, кроме L{n})
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, student_id: int):
        class_id = request.query_params.get("classId")
        if not class_id:
            return Response({"detail": "Укажите параметр classId"}, status=400)

        try:
            classroom = Classroom.objects.select_related("teacher").get(id=class_id)
        except Classroom.DoesNotExist:
            return Response({"detail": "Класс не найден"}, status=404)

        # Проверка прав
        if request.user.role == "STUDENT":
            if request.user.id != int(student_id):
                return Response({"detail": "Студент может смотреть только свой отчёт"}, status=403)
        else:
            # Учитель должен быть владельцем класса
            if request.user.role == "TEACHER" and classroom.teacher_id != request.user.id:
                return Response({"detail": "Доступ запрещён: это не ваш класс"}, status=403)

        # Проверим, что студент вообще относится к классу (состоит в нём напрямую или через команду)
        is_in_class = TeamMembership.objects.filter(team__classroom=classroom, student_id=student_id).exists()
        # Разрешим также, если есть ClassMembership (на случай выдач на класс)
        is_in_class = is_in_class or classroom.memberships.filter(student_id=student_id).exists()
        if not is_in_class:
            return Response({"detail": "Ученик не состоит в указанном классе"}, status=400)

        # Кол-во решённых задач (правильные submissions в рамках этого класса/его команд)
        solved_q = Submission.objects.filter(
            is_correct=True
        ).filter(
            Q(assignment__classroom=classroom) | Q(assignment__team__classroom=classroom)
        ).filter(student_id=student_id)

        solved_count = solved_q.count()

        # Баллы: берём из Score в контексте класса (team=NULL)
        score = Score.objects.filter(student_id=student_id, classroom=classroom, team__isnull=True).first()
        points = score.total_points if score else 0

        # Рейтинг внутри класса: место по total_points
        # (если у кого-то нет Score в классе — считаем 0)
        class_scores = list(
            Score.objects.filter(classroom=classroom, team__isnull=True)
            .values("student_id", "total_points")
        )
        # Добавим текущего, если отсутствует
        if student_id not in [r["student_id"] for r in class_scores]:
            class_scores.append({"student_id": student_id, "total_points": points})
        class_scores.sort(key=lambda r: r["total_points"], reverse=True)

        rank = 1
        for i, r in enumerate(class_scores, start=1):
            if r["student_id"] == int(student_id):
                rank = i
                break

        # Пройденные темы: собираем теги задач, по которым были верные решения
        topics: set[str] = set()
        for row in solved_q.values("assignment__task__tags"):
            tags = row["assignment__task__tags"] or []
            for t in tags:
                if not LEVEL_TAG_RE.match(str(t)):  # исключаем теги L{n}
                    topics.add(str(t))
        topics_list = sorted(topics)

        data = {
            "classId": classroom.id,
            "studentId": int(student_id),
            "solvedCount": solved_count,
            "points": points,
            "rank": rank,
            "topics": topics_list
        }
        return Response(data, status=200)


class ReportExportView(APIView):
    """
    POST /api/reports/export
    Body:
      {
        "scope": "CLASS" | "STUDENT",
        "classId": 123,
        "studentId": 7        # обязательно для scope=STUDENT
      }
    Возвращает Excel (.xlsx) с листами отчётов.
    Доступ:
      - для CLASS: учитель этого класса,
      - для STUDENT: сам студент или учитель класса.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        scope = str(request.data.get("scope", "")).upper()
        class_id = request.data.get("classId")
        student_id = request.data.get("studentId")

        if scope not in {"CLASS", "STUDENT"}:
            return Response({"detail": "scope должен быть CLASS или STUDENT"}, status=400)
        if not class_id:
            return Response({"detail": "Укажите classId"}, status=400)

        try:
            classroom = Classroom.objects.select_related("teacher").get(id=class_id)
        except Classroom.DoesNotExist:
            return Response({"detail": "Класс не найден"}, status=404)

        # Проверка прав
        if scope == "CLASS":
            if request.user.role != "TEACHER" or classroom.teacher_id != request.user.id:
                return Response({"detail": "Доступ запрещён: это не ваш класс"}, status=403)
        else:
            if not student_id:
                return Response({"detail": "Для STUDENT-отчёта нужен studentId"}, status=400)
            student_id = int(student_id)
            if request.user.role == "STUDENT" and request.user.id != student_id:
                return Response({"detail": "Студент может экспортировать только свой отчёт"}, status=403)
            if request.user.role == "TEACHER" and classroom.teacher_id != request.user.id:
                return Response({"detail": "Доступ запрещён: это не ваш класс"}, status=403)

        wb = Workbook()
        wb.remove(wb.active)  # удалим дефолтный лист

        if scope == "CLASS":
            # ----- Лист 1: Сводка по классу -----
            ws = wb.create_sheet("Сводка")
            ws.append(["Класс", classroom.name])
            # Средняя скорость
            correct_subs = Submission.objects.filter(
                is_correct=True
            ).filter(
                Q(assignment__classroom=classroom) | Q(assignment__team__classroom=classroom)
            ).values("checked_at", "assignment__created_at")

            total_td = timedelta(0)
            n = 0
            for r in correct_subs:
                ca, ac = r["checked_at"], r["assignment__created_at"]
                if ca and ac:
                    total_td += (ca - ac)
                    n += 1
            avg_td = (total_td / n) if n else timedelta(0)
            ws.append(["Среднее время решения", _human_timedelta(avg_td)])

            ws.append([])
            ws.append(["ТОП ошибок", "Количество"])
            wrong_rows = Submission.objects.filter(
                is_correct=False
            ).filter(
                Q(assignment__classroom=classroom) | Q(assignment__team__classroom=classroom)
            ).values_list("answer_payload", flat=True)
            counter = Counter(_answer_str(p) for p in wrong_rows if p is not None)
            for value, count in counter.most_common(20):
                ws.append([value, count])

            # ----- Лист 2: Прогресс команд по уровням -----
            ws2 = wb.create_sheet("Прогресс команд")
            ws2.append(["Команда", "Уровень", "Решивших", "Участников", "Завершение, %"])

            teams = list(Team.objects.filter(classroom=classroom).values("id", "name"))
            members_map = {
                row["team_id"]: row["members"]
                for row in TeamMembership.objects.filter(team__classroom=classroom)
                                                 .values("team_id")
                                                 .annotate(members=Count("student_id"))
            }
            rows = Submission.objects.filter(
                is_correct=True,
                assignment__team__classroom=classroom
            ).values("assignment__team_id", "assignment__task__tags", "student_id")

            # считаем уникальные student_id на уровне (team, level)
            seen = set()
            counts = defaultdict(int)  # (team_id, level) -> solved
            for r in rows:
                team_id = r["assignment__team_id"]
                level = _extract_level_from_tags(r["assignment__task__tags"]) or 0
                key = (team_id, level, r["student_id"])
                if key in seen:
                    continue
                seen.add(key)
                counts[(team_id, level)] += 1

            for t in teams:
                tid = t["id"]
                for (team_id, level), solved in sorted(counts.items()):
                    if team_id != tid:
                        continue
                    members = members_map.get(tid, 0)
                    pct = round((solved / max(1, members)) * 100, 1)
                    ws2.append([t["name"], level, solved, members, pct])

            # ----- Лист 3: Рейтинг -----
            ws3 = wb.create_sheet("Рейтинг")
            ws3.append(["Студент", "Очки"])
            class_scores = Score.objects.filter(classroom=classroom, team__isnull=True) \
                                        .select_related("student") \
                                        .order_by("-total_points")
            for s in class_scores:
                ws3.append([s.student.get_username() or s.student_id, s.total_points])

        else:  # scope == "STUDENT"
            # ----- Индивидуальный отчёт -----
            ws = wb.create_sheet("Студент")
            stu = User.objects.filter(id=student_id).first()
            ws.append(["Класс", classroom.name])
            ws.append(["Студент", (stu.full_name or stu.username) if stu else student_id])

            solved_q = Submission.objects.filter(
                is_correct=True, student_id=student_id
            ).filter(
                Q(assignment__classroom=classroom) | Q(assignment__team__classroom=classroom)
            )

            solved_count = solved_q.count()
            score = Score.objects.filter(student_id=student_id, classroom=classroom, team__isnull=True).first()
            points = score.total_points if score else 0

            ws.append(["Решённых задач", solved_count])
            ws.append(["Баллы", points])

            # Ранг
            class_scores = list(
                Score.objects.filter(classroom=classroom, team__isnull=True)
                .values("student_id", "total_points")
            )
            if student_id not in [r["student_id"] for r in class_scores]:
                class_scores.append({"student_id": student_id, "total_points": points})
            class_scores.sort(key=lambda r: r["total_points"], reverse=True)
            rank = 1
            for i, r in enumerate(class_scores, start=1):
                if r["student_id"] == int(student_id):
                    rank = i
                    break
            ws.append(["Позиция в рейтинге", rank])

            # Темы (теги, кроме L{n})
            topics: set[str] = set()
            for row in solved_q.values("assignment__task__tags"):
                for t in (row["assignment__task__tags"] or []):
                    if not LEVEL_TAG_RE.match(str(t)):
                        topics.add(str(t))
            ws.append([])
            ws.append(["Пройденные темы"])
            for t in sorted(topics):
                ws.append([t])

            # Подробный список решённых задач
            ws2 = wb.create_sheet("Задачи")
            ws2.append(["Название", "Уровень", "Получено очков", "Дата проверки"])
            for r in Submission.objects.filter(
                is_correct=True, student_id=student_id
            ).filter(
                Q(assignment__classroom=classroom) | Q(assignment__team__classroom=classroom)
            ).select_related("assignment__task").order_by("-checked_at"):
                task = r.assignment.task
                level = _extract_level_from_tags(task.tags) or ""
                ws2.append([task.title, level, r.points_awarded, r.checked_at.strftime("%Y-%m-%d %H:%M") if r.checked_at else ""])

        # Отдаём Excel
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"report_{scope.lower()}_{class_id}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
